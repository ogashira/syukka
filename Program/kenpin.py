#! python
# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import re
import csv
import barcode
import platform
import pickle
from barcode.writer import ImageWriter
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import unicodedata



class Kenpin(object):

    def __init__(self, factory, packingHinban, untinForUriage, myfolder):
        # ﾃｽﾄ用にmyfolderが'./'だったらｶﾚﾝﾄﾃﾞｨﾚｸﾄﾘに入れる
        if factory == 'toke':
            if myfolder == './':
                self.kenpin_folder = r'./kenpin_toke.csv'
            else:
                self.kenpin_folder = (r'\\192.168.3.204\effitA_HT\送信データ'
                                                                 r'\kenpin.csv')
            self.syukka_koujou = '出荷工場：@0002 土気工場'
            self.factory = '土気'
        elif factory == 'honsya':
            if myfolder == './':
                self.kenpin_folder = r'./kenpin_honsya.csv'
            else:
                self.kenpin_folder = r'C:\effitA_HT\送信データ\kenpin.csv'

            self.syukka_koujou = '出荷工場：@0001 本社工場'
            self.factory = '本社'
        else:
            self.kenpin_folder = r'./kenpin_mac.csv'
            self.syukka_koujou = 'mac : @0000 fromMac'
            self.factory = 'Mac'

        self.packingHinban = packingHinban

        self.untinForUriage = untinForUriage

        self.myfolder = myfolder




    def get_kenpin(self):
        
        # data.pickleからﾀﾞｳﾝﾛｰﾄﾞ
        with open(r'./data.pickle', 'rb') as f:
            data_loaded = pickle.load(f)

        unsou_dic2 = data_loaded['unsou_dic2'] 
        haisou_kubun = data_loaded['haisou_kubun']

        def get_kubun(df_row):
            iraisaki = df_row['依頼先']
            yotei_souko = df_row['出荷予定倉庫_y']
            syukkabi = df_row['出荷予定日']
            cans = df_row['cans']
            unsou = unsou_dic2[iraisaki][0]
            unsou_code = unsou_dic2[iraisaki][1]
            if '営業所' in yotei_souko:
                kubun = '営業所'
                kubun_no = 3
            elif '土曜配達' in yotei_souko:
                kubun = '土曜配達'
                kubun_no = 2
            elif '曜日' in yotei_souko:
                kubun = '曜日違い'
                kubun_no = 4
            elif '祝日配達' in yotei_souko:
                kubun = '祝日配達'
                kubun_no = 5
            else:
                kubun = '通常'
                kubun_no = 1
                
            cans = int(cans)
            syukkabi = syukkabi.replace('/', '')
            
            return pd.Series([unsou, unsou_code, kubun, kubun_no, cans,
                            syukkabi]
                            )


        merg_data = self.packingHinban[
        ['受注ＮＯ','受注行ＮＯ','出荷予定倉庫', '品名','受注単位', 
        '納入先名称１']
        ]

        kenpin = pd.merge(self.untinForUriage, merg_data, 
                on=['受注ＮＯ','受注行ＮＯ'], how = 'left')
        
        kenpin[['unsou','unsou_code','kubun','kubun_no','cans',
            '出荷予定日']]  = kenpin.apply(get_kubun, axis = 1)

        kenpin = kenpin[
        ['得意先コード','納入先コード','unsou_code','unsou','kubun_no','kubun',
        '出荷予定日','hinban','品名','lot','cans','受注数量','受注単位',
        '納入先名称１','輸出向先', '得意先注文ＮＯ', '備考','add']
        ]

        # lotが辞書で２つ以上のものをリテラルにして、行を増やす。>>>>>>>>>>>>>>>
        
        # df_listにkenpinを１行ずつ入れる
        df_list = []
        for i in range(len(kenpin)):
            df = kenpin.iloc[[i]]
            df_list.append(df)
        
        # lotのkeyとvalueをcansと受注数量に入れながら、ｺﾋﾟｰを
        # df_list2に入れていく
        df_list2 = []    
        for df in df_list:
            lots = df.iloc[0,9]
            if lots == {}:
                lots = {'short':0}
            for k, v in lots.items():
                df2 = df.copy()
                cans = df2.loc[:, 'cans']
                kg = df2.loc[:,'受注数量']
                ratio = kg/cans
                # lotが２種類以上あった場合のcansとkgを計算する。
                df2.loc[:,'lot'] = k
                if k != 'short':
                    df2.loc[:, 'cans'] = int(v)
                    df2.loc[:, '受注数量'] = int(v * ratio)
                else:
                    df2.loc[:,'lot'] = None
                    
                df_list2.append(df2)
        
        df_col = df_list2[0].columns
        
        df_kara = pd.DataFrame(index=[], columns=df_col)
        for line in df_list2:
            df_kara = pd.concat([df_kara, line])
        
        
        df_kara2 = df_kara.sort_values(['unsou_code', 'kubun_no'])

        return df_kara2


    def create_kenpin(self):

        df_kara2 = self.get_kenpin()
        df_kara2 = df_kara2.loc[(df_kara2['輸出向先'] != 'y') & 
                                (df_kara2['hinban'] != '999998'), :]
        df_kara2 = df_kara2[['unsou_code','unsou','kubun_no','kubun',
                            '出荷予定日','hinban','品名','lot','cans','受注数量']]
        
        try:
            df_kara2.to_csv(self.kenpin_folder, index=False, header = False , 
                            encoding='cp932')
        except Exception as ex:
            print('************kenpin.csv作成エラー****************')
            print('Folderが見つからないので、{}のkenpin.csvをmyfolderに放り込みます'
                .format(self.factory))
            df_kara2.to_csv(self.myfolder + '/kenpin_' + self.factory + '.csv', 
                index=False, header = False , encoding='cp932')





    def get_syukka_jisseki_syoukai(self):
        

        # 相手先略称ﾃﾞｰﾀの取得
        pf = platform.system()
        
        if pf == 'Windows':
            mypath = r'//192.168.1.247/共有/受注check/master/order_nounyuusaki.csv'
        elif pf == 'Linux':
            mypath = r'/mnt/public/受注check/master/order_nounyuusaki.csv'
        else:
            mypath = r'../master/selfMade/order_nounyuusaki.csv'
            
        nounyuusaki = pd.read_csv(mypath, encoding = 'cp932')

        #merge用データに加工する
        merge_data = nounyuusaki[['相手先コード１','相手先コード２','相手先略称']]
        merge_data = merge_data.drop_duplicates(['相手先コード１','相手先コード２'])
        merge_data = merge_data.rename(
            columns = {'相手先コード１':'得意先コード', '相手先コード２':'納入先コード',
                       '相手先略称':'納入先名'}
                    )

        merge_data = merge_data.fillna({'納入先コード':'noData'})

        # merge_dataのNaNを空文字にしておく。こうしないとmergeができない。
        merge_data = merge_data.fillna('')
        
        kenpin_moto = self.get_kenpin()
        kenpin_moto = kenpin_moto.fillna('')
        
        kenpin_merge = pd.merge(kenpin_moto, merge_data, 
                                on = ['得意先コード', '納入先コード'], how = 'left')

        # unsou_codeとkubun_noと輸出向先のタプルをsetに入れて重複をなくす。>>>>>>>>>>>>>
        # {(U0009, 1, ''), (U0009, 1, 'y'), (U0005, 1, '')......}
        unsou_set = set()
        for i in range(len(kenpin_merge)):
            unsou_code = kenpin_merge.iloc[i, 2]
            kubun_no = kenpin_merge.iloc[i, 4]
            yusyutu = kenpin_merge.iloc[i, 14]

            t = (unsou_code, kubun_no, yusyutu)
            unsou_set.add(t)
        # <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

        i = 0
        wb = openpyxl.Workbook()
        for unsou_code, kubun_no, yusyutu in unsou_set:

            kenpin_split = kenpin_merge.loc[
                                     (kenpin_merge['unsou_code']== unsou_code) &
                                     (kenpin_merge['kubun_no'] == kubun_no) &
                                     (kenpin_merge['輸出向先'] == yusyutu), :]
            unsou = kenpin_split.iloc[0,3]
            kubun = kenpin_split.iloc[0,5]
            if kenpin_split.iloc[0, 14] == '':
                yusyutu = 'N'
            else:
                yusyutu = 'Y'

            unsou_gyousya = '運送業者：{} {}   配送区分：{} {}  輸出：{}' \
                .format(unsou_code, unsou, kubun_no, kubun, yusyutu)

            sheet_name = '{}_{}_輸出{}'.format(unsou, kubun, yusyutu)
            barcode_str = unsou_code + str(kubun_no)

            # 一旦バーコードをpngで保存 
            code39 = barcode.get_barcode_class('code39')
            barcode_img = code39(barcode_str, writer=ImageWriter(), add_checksum = False)
            barcode_img.save(r'./barcode_{}'.format(i))
            
            ws_new = wb.create_sheet(title = sheet_name,index= i)
            ws = wb[sheet_name]

            ws.sheet_view.showGridLines = False  # 枠線を消す
            ws['E1'].value = '出荷実績照会'
            ws['A2'].value = self.syukka_koujou
            ws['A3'].value = unsou_gyousya

            

            kenpin_split = kenpin_split[['出荷予定日','納入先名','品名','lot','受注数量',
                '受注単位','得意先注文ＮＯ','cans','備考','輸出向先','add']]
            syukka_jisseki_syoukai = kenpin_split.rename(
            columns = {'出荷予定日':'売上日','lot':'ﾛｯﾄNo.','受注数量':'売上数量',
            '受注単位':'単位','得意先注文ＮＯ':'注文No.','cans':'缶数'}
            )

            side = Side(style='thin', color='000000')
            border = Border(top=side, bottom=side, left=side, right=side)

            ws.cell(4 , 1).value = '行'
            ws.cell(4 , 1).border = border
            ws.cell(4 , 2).value = '売上日'
            ws.cell(4 , 2).border = border
            ws.cell(4 , 3).value = '納入先名'
            ws.cell(4 , 3).border = border
            ws.cell(4 , 4).value = '品名'
            ws.cell(4 , 4).border = border
            ws.cell(4 , 5).value = 'ﾛｯﾄNo.'
            ws.cell(4 , 5).border= border
            ws.cell(4 , 6).value = '売上数量'
            ws.cell(4 , 6).border = border
            ws.cell(4 , 7).value = '単位'
            ws.cell(4 , 7).border = border
            ws.cell(4 , 8).value = '注文No.'
            ws.cell(4 , 8).border = border
            ws.cell(4 , 9).value = '缶数'
            ws.cell(4 , 9).border = border
            ws.cell(4 , 10).value = '備考'
            ws.cell(4 , 10).border = border
            ws.cell(4 , 11).value = 'y'
            ws.cell(4 , 11).border = border
            ws.cell(4 , 12).value = 'add'
            ws.cell(4 , 12).border = border



            for j in range(len(syukka_jisseki_syoukai)):
                for k in range(len(syukka_jisseki_syoukai.columns)):
                    ws.cell(j+5, 1).value = j + 1
                    ws.cell(j+5, 1).border = border 
                    ws.cell(j+5, k+2).value = syukka_jisseki_syoukai.iloc[j, k]
                    ws.cell(j+5, k+2).border = border
            
            
            def get_east_asian_width_count(text):
                # 全角英数は'F',全角かなは'W', 特殊文字は'A'が返る。 
                count = 0
                for c in text:
                    if unicodedata.east_asian_width(c) in 'FWA':
                        count += 1.7
                    else:
                        count += 1
                return count


            # columnの文字数から列を調整する
            for col in ws.columns:
                max_length = 0
                column = col[0].column
            
                for cell in col:
                    if cell.row < 4: # ３行目まではautofitさせない。
                        continue
                    cell.font = Font(size=10)
                    count = get_east_asian_width_count(str(cell.value))

                    if count > max_length:
                        max_length = count
            
                adjusted_width = (max_length + 0.5) * 1.1
                ws.column_dimensions[get_column_letter(column)].width = adjusted_width
                # ws.column_dimensions[column].width = adjusted_width

            # wsにバーコードを貼り付け
            img = barcode_img
            img = openpyxl.drawing.image.Image(r'./barcode_{}.png'.format(i))
            img.width = 72*3.5
            img.height = 25*2.5
            ws.add_image(img, 'H1')
            
            ws.row_dimensions[1].height = 23
            for j in range(2, ws.max_row + 1):
                ws.row_dimensions[j].height = 18

             
            # 印刷設定
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_margins.left = 0.2
            ws.page_margins.right = 0.2
            ws.page_margins.top = 0.8
            ws.page_margins.bottom = 0.8

            i += 1

        wb.save('{}/出荷実績照会_{}.xlsx'.format(self.myfolder, self.factory))

        
        '''
        2020/12/18 中村さんからの依頼
        出荷実績照会のｹｲﾋﾝ便のみ、国内と輸出でｼｰﾄを分ける
        輸出の場合は最終列に'y'が入っているからそれで識別する
        '''
        

