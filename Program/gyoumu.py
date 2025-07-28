#! python
# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl
import unicodedata
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.utils import get_column_letter
import jaconv
from recorder import *

class Gyoumu:

    def __init__ (self, myfolder):
        self.myfolder = myfolder



    def get_sorting(self, PH, myfolder, factory):
        """
        PHはmodi_PHのこと
        """

        def get_dupli(packingHinban):

            '''輸出と国内にわけてからdupli後concatする
            yがない場合空のdataframeが出来、それをconcatすると、
            dupliのbooleanがfloatになってしまう。True→1, False→0
            そこで、concatの後に、dupliのﾃﾞｰﾀをboolにキャストする。'''
        
            
            def moji_henkan(moji): 
                moji = jaconv.z2h(moji,digit=True,ascii=True) 
                moji = moji.replace(' ', '')
                moji = moji.replace('　', '')
                moji = moji.upper()
                return moji


            PH_domestic = packingHinban.loc[packingHinban['輸出向先'] \
                != 'y',:].copy()
            PH_export = packingHinban.loc[packingHinban['輸出向先'] \
                == 'y', :].copy()
            
            #国内のdupli 備考内の文字を半角小文字空白無しにする。
            PH_domestic.loc[:,'mojiHenkan'] = PH_domestic['備考'].map(moji_henkan)
            PH_domestic = PH_domestic.sort_values(by=['納入先名称１', 'mojiHenkan'])

            # 2024/11/5 20241024の出荷で,南部化成向けと川真向けが同じになってしまった
            # ﾊﾞｸﾞを修正->duplicatedの要素に得意先コードを追加した
            PH_domestic['dupli'] = PH_domestic.duplicated(
                               subset=['得意先コード', '納入先名称１', 'mojiHenkan'],keep='first')

            #輸出のdupli
            PH_export = PH_export.sort_values(by='得意先注文ＮＯ')
            PH_export['dupli'] = PH_export.duplicated(subset=['得意先注文ＮＯ'] \
                                                      ,keep='first')
            
            #concat後indexをreset
            PH_con = pd.concat([PH_domestic, PH_export], sort=True)
            PH_con.loc[:,'dupli'] = PH_con.loc[:, 'dupli'].map(lambda x: bool(x))
            # 下の２回のreset_indexによって昇順の'index'列を作る
            PH_con.reset_index(inplace = True, drop=True)
            PH_con.reset_index(inplace = True)
            

            # 大ﾊﾞｸﾞ→このsort_valuesによって、輸出のpackingの区分けがめちゃlめ
            #ちゃになる。直した得意先コードと昇順index列でソート
            PH_con = PH_con.sort_values(by = ['得意先コード', 'index'])

            PH_con = PH_con.drop('index', axis = 1) # index列を削除
            
            PH_con.reset_index(inplace=True, drop=True)
            
            
            return PH_con


        packingHinban = PH.sort_values(['得意先コード', '備考'])


        #packingHinbanにdupliを追加(dupliは仕分けの目安、重複の目印)
        packingHinban = get_dupli(packingHinban)


        '''
        dataframeの行を１行ずつ分解して、listに入れる。
        分解したdataframeのindexはそれぞれ違うことに注意する。
        index １つ目は0、２つ目は1、３つ目は２.....となっている。
        '''
        df_list = []
        for i in range(len(packingHinban)):
            #copyしたものをappendしないと後々連鎖代入となってwarningがでる。→※1
            # ※1 で、df_list2[i]をcopyして連鎖代入を避けることもできる
            df = packingHinban.iloc[i : i + 1].copy()
            df_list.append(df)


        '''
        duplicatedのdfのkeep = 'first'にしたので、最初のdupliは必ずfalseになる。
        dupliがfalseだったら下のjのroopにはいる。roopでは、下のdupliがtrueならば、
        dfをconcatし続ける。
        falseが出た時点でroopを抜けてiのroopを続ける。
        iのroopは、dupliがtrueの時は何もしない。
        '''
        df_list2 = df_list.copy()
        if len(df_list2) == 1:
            df_list2[0].loc[:,'総重量'] = df_list2[0].loc[:,'weight'].sum()
        else:        
            for i in range(len(df_list2)):
                #indexがそれぞれ違うので、loc[i, 'supli'] にする
                if df_list2[i].loc[i, 'dupli'] == False:
                    j = i + 1
                    for j in range(j,len(df_list2)):
                        if df_list2[j].loc[j,'dupli'] == True:
                            df_list2[i] = pd.concat([df_list2[i], df_list2[j]])
                            df_list2[i]['dupli'] = df_list2[i]['dupli'].astype(object)
                            df_list2[i].loc[:,'dupli'] = 'concat'
                        else:
                            break
                    # ※1 ここでwarningがでる
                    df_list2[i].loc[:,'総重量'] = df_list2[i].loc[:,'weight'].sum()


        #dupliがFalseまたはconcatのものだけをlist3に入れる。
        df_list3 = []
        
        for i in range(len(df_list2)):
            if df_list2[i].loc[i, 'dupli'] != True:
                df_list3.append(df_list2[i])


        # df_col = df_list3[0].columns
        df_col = packingHinban.columns

        df_kara = pd.DataFrame(index=[], columns=df_col)
        df_kara.loc['a'] = '<' * 5

        result = df_kara.copy()
        for line in df_list3:
            result = pd.concat([result, line], sort = True)
            result = pd.concat([result, df_kara], sort = True)



        result = result[['依頼先','cans','総重量','得意先コード','納入先コード',
                         '納入先名称１', '品名','得意先注文ＮＯ','備考','納期',
                         '出荷','出荷予定倉庫','add']]




        recorder = Recorder(self.myfolder)
        
        txt = '業務出荷準備用ﾃﾞｰﾀ（{}）'.format(factory)
        recorder.out_log('')
        recorder.out_file('')
        recorder.out_log(txt, '\n')
        recorder.out_log(result, '\n')
        recorder.out_file(txt, '\n')
        recorder.out_file(result, '\n')

        #packingHinbanをexcel形式でfilePathに保存する。
        filePath = '{}/{}業務_packing.xlsx'.format(myfolder, factory)
        result.to_excel(filePath, index = False)
        #recorder.out_csv(result, filePath)

        del recorder


        return result


    def get_excel_style(self, filePath):
        wb =  openpyxl.load_workbook(filePath)
        ws = wb['Sheet1']
        
        max_col = ws.max_column
        max_row = ws.max_row

        def get_east_asian_width_count(text):
            # 全角英数は'F',全角かなは'W', 特殊文字は'A'が返る。 
            count = 0
            for c in text:
                if unicodedata.east_asian_width(c) in 'FWA':
                    count += 1.7
                else:
                    count += 1
            return count

        i = 2
        while True:
            if ws.cell(i, 1).value == None:
                break
            if ws.cell(i, 1).value == '<<<<<':
                ws.delete_rows(i)
                
                for j in range(1, max_col+1):
                    border = Border(top = Side(style='medium', color='000000'))
                    ws.cell(i, j).border = border
                
                # １行削除した分だけ-1する
                i -= 1 

            i += 1

        # columnの文字数から列を調整する
        for col in ws.columns:
            max_length = 0
            column = col[0].column
        
            for cell in col:
                if cell.row == 1:
                    continue
                cell.font = Font(size=10)
                count = get_east_asian_width_count(str(cell.value))

                if count > max_length:
                    max_length = count
        
            adjusted_width = (max_length + 0) * 1.1
            ws.column_dimensions[get_column_letter(column)].width = adjusted_width
            # ws.column_dimensions[column].width = adjusted_width
            


        for i in range(max_row + 1):
            ws.row_dimensions[i].height = 20

        # 印刷設定
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.8
        ws.page_margins.bottom = 0.8
            
        wb.save(filePath)

        
